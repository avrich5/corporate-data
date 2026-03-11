import warnings
from pathlib import Path

import pandas as pd
import openpyxl


def read_all_sheets(path: Path) -> dict[str, pd.DataFrame]:
    """
    Reads all sheets from an Excel file using pandas, with header=None.
    Suppresses openpyxl warnings that are expected (e.g. Data Validation extension warnings).
    
    Args:
        path: Path to the Excel file (should be standard relative path usage in main).
        
    Returns:
        A dictionary mapping sheet names to their raw DataFrames.
        
    Raises:
        FileNotFoundError: If the specified Excel file is not found.
    """
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found at: {path}")

    # Suppress openpyxl warnings specifically during read
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=UserWarning)
        # We read all sheets blindly and don't assume a header to allow stage 1 detection
        return pd.read_excel(path, sheet_name=None, header=None)


def get_merged_cell_ranges(path: Path, sheet_name: str) -> list[str]:
    """
    Extracts a list of merged cell ranges from a specific sheet in an Excel file.
    
    Args:
        path: Path to the Excel file.
        sheet_name: Name of the sheet to inspect.
        
    Returns:
        A list of string coordinate ranges, e.g., ["A1:C1", "D2:D4"].
        
    Raises:
        FileNotFoundError: If the specified Excel file is not found.
    """
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found at: {path}")

    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=UserWarning)
        # load_workbook read_only=False is required to access merged_cells
        # Keep data_only=True so we don't process formula nodes unnessarily
        wb = openpyxl.load_workbook(path, data_only=True)
        
    if sheet_name not in wb.sheetnames:
        return []
        
    ws = wb[sheet_name]
    
    # openpyxl returns MultiCellRange, we extract string representations
    return [str(rng) for rng in ws.merged_cells.ranges]
